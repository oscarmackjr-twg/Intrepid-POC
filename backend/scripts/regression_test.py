"""Data regression test harness for pipeline output validation.

Discovers buy-date test case folders and runs all four pipeline phases against
each, diffs generated outputs against expected outputs byte-for-byte, and prints
a summary report.  Optionally writes an Excel report following the TestMatrix
template format.

Phases run per test case:
  1. Pre-Funding  — scripts/run_pipeline_cli.py (validates eligibility checks)
  2. Tagging      — scripts/tagging.py (splits exhibit files into _sg/_cibc)
  3. Funding SG   — scripts/final_funding_sg.py
  4. Funding CIBC — scripts/final_funding_cibc.py
  5. CashFlow     — cashflow/compute/run_purchase_package.py (SG and CIBC)

Phases 2-5 run against a temp copy of files_required/ so the golden TestData
directory is never modified.  Generated outputs land in temp_dir/output/ and
temp_dir/output_share/, which are then diffed against the golden
test_case_dir/output/ and test_case_dir/output_share/.

Usage (from repo root):
    python backend/scripts/regression_test.py
    python backend/scripts/regression_test.py --test-data C:\\Users\\omack\\Downloads\\TestData
    python backend/scripts/regression_test.py --pdate 2026-02-24 --tday 2026-02-19
    python backend/scripts/regression_test.py --no-cleanup  # keep temp dirs for inspection
    python backend/scripts/regression_test.py --skip-phase1  # skip pre-funding (faster)
    python backend/scripts/regression_test.py --report report.xlsx  # custom report path

TestData folder structure:
    <test-data-dir>/
        {buy_date_folder}/          # e.g. 93rd_buy or any name
            files_required/         # all pipeline input + intermediate files
            output/
                <expected output files>
            output_share/           # optional
                <expected output_share files>
        dates.json                  # optional per-folder date config

Exit code: 0 if all cases PASS, 1 if any case FAILS.
"""
import argparse
import filecmp
import fnmatch
import json
import os
import shutil
import subprocess
import sys
import tempfile
from datetime import date, datetime, timedelta
from pathlib import Path


# ---------------------------------------------------------------------------
# Path helpers
# ---------------------------------------------------------------------------

def _repo_root() -> Path:
    """Return the repo root (parent of backend/)."""
    return Path(__file__).resolve().parent.parent.parent


def _default_backend_dir() -> Path:
    return _repo_root() / "backend"


# ---------------------------------------------------------------------------
# Discovery
# ---------------------------------------------------------------------------

def discover_test_cases(test_data_dir: Path) -> list[Path]:
    """Return sorted list of valid test-case directories under test_data_dir."""
    if not test_data_dir.exists():
        print(f"[WARN] Test data directory not found: {test_data_dir}", file=sys.stderr)
        return []

    cases = []
    for subdir in sorted(test_data_dir.iterdir()):
        if not subdir.is_dir():
            continue
        files_required = subdir / "files_required"
        has_files = files_required.is_dir() and any(f.is_file() for f in files_required.iterdir())
        has_expected = (subdir / "output").is_dir() or (subdir / "output_share").is_dir()
        if not has_files:
            print(f"[WARN] Skipping {subdir.name}: no input files found in files_required/")
            continue
        if not has_expected:
            print(f"[WARN] Skipping {subdir.name}: no output/ or output_share/ directory")
            continue
        cases.append(subdir)

    return cases


# ---------------------------------------------------------------------------
# Date helpers
# ---------------------------------------------------------------------------

def _load_dates_config(test_data_dir: Path) -> dict:
    """Load optional dates.json from TestData root. Returns {folder_name: {pdate, tday}}."""
    config_path = test_data_dir / "dates.json"
    if not config_path.exists():
        return {}
    try:
        with open(config_path) as f:
            return json.load(f)
    except Exception as exc:
        print(f"[WARN] Could not read dates.json: {exc}")
        return {}


def _derive_dates(
    folder_name: str,
    cli_pdate: str | None,
    cli_tday: str | None,
    dates_config: dict | None = None,
) -> tuple[str, str]:
    """Return (pdate, tday) for a test case."""
    today_str = date.today().isoformat()

    if cli_pdate and cli_tday:
        return cli_pdate, cli_tday

    if dates_config and folder_name in dates_config:
        entry = dates_config[folder_name]
        pdate = cli_pdate or entry.get("pdate", folder_name)
        tday = cli_tday or entry.get("tday", today_str)
        return pdate, tday

    try:
        datetime.strptime(folder_name, "%Y-%m-%d")
        pdate = folder_name
    except ValueError:
        pdate = folder_name

    tday = cli_tday if cli_tday else today_str
    pdate = cli_pdate if cli_pdate else pdate
    return pdate, tday


def _derive_date_vars(tday: str) -> dict:
    """Derive date variables needed by final_funding scripts from tday (YYYY-MM-DD)."""
    tday_dt = datetime.strptime(tday, "%Y-%m-%d")
    yesterday_dt = tday_dt - timedelta(days=1)
    curr_date = tday_dt.strftime("%m-%d-%Y")
    yesterday_str = yesterday_dt.strftime("%m-%d-%Y")
    first_of_month = tday_dt.replace(day=1)
    last_of_prev_month = first_of_month - timedelta(days=1)
    last_end = f"{last_of_prev_month.year}_{last_of_prev_month.month:03}_{last_of_prev_month.day:02}"
    fd = first_of_month.strftime("%Y-%m-%d")
    return {
        "curr_date": curr_date,
        "yesterday": yesterday_str,
        "last_end": last_end,
        "fd": fd,
    }


# ---------------------------------------------------------------------------
# Phase classification helpers (used by report generator)
# ---------------------------------------------------------------------------

# Column order for Excel report phases
PHASE_COLS = [
    "Pre-Funding\n(AllInOne)",
    "Tagging",
    "Final Funding\nSG",
    "Final Funding\nCIBC",
    "CashFlow",
]
PHASE_KEYS = ["prefunding", "tagging", "funding_sg", "funding_cibc", "cashflow"]


def _classify_output_file(filename: str) -> str:
    """Return PHASE_KEYS entry for an output file based on its name."""
    name = Path(filename).name
    nl = name.lower()

    # CashFlow
    for pat in ["sfc_cashflows_*", "twg_cashflows_*", "loans_data_*",
                "cashflows_*", "cashflow profile*"]:
        if fnmatch.fnmatch(nl, pat):
            return "cashflow"

    # Final Funding SG-specific
    for pat in ["everyloan_sg*", "borrowing_file_sg*", "concentration_final_sg*",
                "flagged_loans_*_sg*", "notes_flagged_loans_*_sg*"]:
        if fnmatch.fnmatch(nl, pat):
            return "funding_sg"

    # Final Funding CIBC-specific
    for pat in ["everyloan_cibc*", "borrowing_file_cibc*", "concentration_final_cibc*",
                "flagged_loans_*_cibc*", "notes_flagged_loans_*_cibc*"]:
        if fnmatch.fnmatch(nl, pat):
            return "funding_cibc"

    # Tagging outputs (split exhibit files written to output/)
    for pat in ["tagging_summary*", "tagging_allocation*"]:
        if fnmatch.fnmatch(nl, pat):
            return "tagging"

    # Final Funding shared (comap, flags, purchase price)
    for pat in ["comap_not_passed*", "purchase_price_mismatch*",
                "flagged_loans_*", "notes_flagged_loans_*"]:
        if fnmatch.fnmatch(nl, pat):
            return "funding_sg"

    return "funding_sg"  # default


def _classify_input_file(filename: str) -> str:
    """Return PHASE_KEYS entry for an input file (which phase introduces it)."""
    name = Path(filename).name
    nl = name.lower()

    # Tagging outputs (split exhibit files)
    if ("_sg" in nl or "_cibc" in nl) and ("exhibit" in nl or "fx3_" in nl):
        return "tagging"

    # Pre-Funding outputs (raw exhibit files generated by AllInOne)
    if ("exhibitatoformofsalenotice" in nl.replace(" ", "").replace("-", "")
            and "_sg" not in nl and "_cibc" not in nl
            and "pre-funding" not in nl):
        return "prefunding"

    # Pre-Funding inputs (SFC / SFTP)
    for pat in ["tape20loans_*", "sfy_*pre-funding*", "prime_*pre-funding*",
                "fx3 - twg*", "fx4 - twg*", "funding request*"]:
        if fnmatch.fnmatch(nl, pat):
            return "prefunding"

    # Monthly servicing files
    if fnmatch.fnmatch(nl, "fx3_2*") or fnmatch.fnmatch(nl, "fx4_2*"):
        return "prefunding"

    return ""  # baseline/fixed files


# ---------------------------------------------------------------------------
# Diff logic
# ---------------------------------------------------------------------------

def _collect_files(directory: Path) -> dict[str, Path]:
    """Return {relative_path_str: absolute_path} for all files recursively."""
    result = {}
    if not directory.exists():
        return result
    for f in directory.rglob("*"):
        if f.is_file():
            rel = str(f.relative_to(directory))
            result[rel] = f
    return result


def _diff_directories(
    generated_dir: Path,
    expected_dir: Path,
    label: str,
) -> tuple[list[str], list[str], list[str]]:
    """Compare two directories. Returns (diffs, missing, extras)."""
    diffs: list[str] = []
    missing: list[str] = []
    extras: list[str] = []

    if not expected_dir.exists():
        return diffs, missing, extras

    expected_files = _collect_files(expected_dir)
    generated_files = _collect_files(generated_dir) if generated_dir.exists() else {}

    for rel, exp_path in expected_files.items():
        display = f"{label}/{rel}"
        if rel not in generated_files:
            missing.append(display)
        else:
            try:
                same = filecmp.cmp(str(generated_files[rel]), str(exp_path), shallow=False)
            except OSError as exc:
                diffs.append(f"{display} (compare error: {exc})")
                continue
            if not same:
                diffs.append(display)

    for rel in generated_files:
        if rel not in expected_files:
            extras.append(f"{label}/{rel}")

    return diffs, missing, extras


# ---------------------------------------------------------------------------
# Phase runners
# ---------------------------------------------------------------------------

def _run_phase(
    phase_name: str,
    cmd: list[str],
    cwd: str,
    env: dict | None = None,
    timeout: int = 300,
) -> tuple[bool, str | None]:
    """Run a subprocess for a pipeline phase. Returns (ok, error_msg)."""
    run_env = os.environ.copy()
    if env:
        run_env.update(env)

    try:
        proc = subprocess.run(
            cmd,
            cwd=cwd,
            capture_output=True,
            text=True,
            timeout=timeout,
            env=run_env,
        )
        if proc.stdout:
            for line in proc.stdout.splitlines():
                print(f"  [{phase_name}] {line}")
        if proc.stderr:
            for line in proc.stderr.splitlines():
                print(f"  [{phase_name} STDERR] {line}")

        if proc.returncode != 0:
            return False, f"{phase_name} exited with code {proc.returncode}"
        return True, None

    except subprocess.TimeoutExpired:
        return False, f"{phase_name} timed out after {timeout}s"
    except Exception as exc:
        return False, f"{phase_name} exception: {exc}"


# ---------------------------------------------------------------------------
# Per-test-case runner
# ---------------------------------------------------------------------------

def run_test_case(
    test_case_dir: Path,
    backend_dir: Path,
    cli_pdate: str | None,
    cli_tday: str | None,
    no_cleanup: bool,
    skip_phase1: bool,
    dates_config: dict | None = None,
) -> dict:
    """Run all pipeline phases for one test case. Returns result dict."""
    pdate, tday = _derive_dates(test_case_dir.name, cli_pdate, cli_tday, dates_config)
    date_vars = _derive_date_vars(tday)
    buy_num = test_case_dir.name.replace("_buy", "")

    result = {
        "name": test_case_dir.name,
        "test_case_dir": test_case_dir,
        "pdate": pdate,
        "tday": tday,
        "buy_num": buy_num,
        "date_vars": date_vars,
        "status": "FAILED",
        "phases_ok": [],
        "phases_failed": [],
        "diffs": [],
        "missing": [],
        "extra": [],
        "error": None,
        "work_dir": None,
    }

    print(f"\n[BUY DATE: {test_case_dir.name}]")
    print(f"  pdate={pdate}  tday={tday}  curr_date={date_vars['curr_date']}  buy_num={buy_num}")

    # Phase 1: Pre-Funding
    if not skip_phase1:
        print("  [Phase 1] Pre-Funding...")
        ok, err = _run_phase(
            "Phase1",
            [
                sys.executable,
                "scripts/run_pipeline_cli.py",
                "--folder", str(test_case_dir),
                "--pdate", pdate,
                "--tday", tday,
            ],
            cwd=str(backend_dir),
            timeout=300,
        )
        if ok:
            result["phases_ok"].append("Phase1:PreFunding")
            print("  [Phase 1] OK")
        else:
            result["phases_failed"].append(f"Phase1:PreFunding — {err}")
            result["error"] = err
            print(f"  [Phase 1] FAILED: {err}")
            return result
    else:
        print("  [Phase 1] Pre-Funding skipped (--skip-phase1)")
        result["phases_ok"].append("Phase1:PreFunding(skipped)")

    # Set up temp work dir for Phases 2-5
    tmp_parent = Path(tempfile.mkdtemp(prefix="regression_test_"))
    work_dir = tmp_parent / "work"
    work_dir.mkdir()
    result["work_dir"] = work_dir

    try:
        shutil.copytree(str(test_case_dir / "files_required"), str(work_dir / "files_required"))
        (work_dir / "output").mkdir()
        (work_dir / "output_share").mkdir()
        print(f"  Work dir: {work_dir}")

        funding_env = {
            "FOLDER": str(work_dir),
            "PDATE": pdate,
            "CURR_DATE": date_vars["curr_date"],
            "YESTERDAY": date_vars["yesterday"],
            "LAST_END": date_vars["last_end"],
            "FD": date_vars["fd"],
            "BUY_NUM": buy_num,
        }

        # Phase 2: Tagging
        print("  [Phase 2] Tagging...")
        ok, err = _run_phase(
            "Phase2",
            [sys.executable, "scripts/tagging.py"],
            cwd=str(backend_dir),
            env={"FOLDER": str(work_dir), "PDATE": pdate},
            timeout=120,
        )
        if ok:
            result["phases_ok"].append("Phase2:Tagging")
            print("  [Phase 2] OK")
        else:
            result["phases_failed"].append(f"Phase2:Tagging — {err}")
            result["error"] = err
            print(f"  [Phase 2] FAILED: {err}")
            return result

        # Phase 3a: Funding SG
        print("  [Phase 3a] Funding SG...")
        ok, err = _run_phase(
            "Phase3-SG",
            [sys.executable, "scripts/final_funding_sg.py"],
            cwd=str(backend_dir),
            env=funding_env,
            timeout=300,
        )
        if ok:
            result["phases_ok"].append("Phase3a:FundingSG")
            print("  [Phase 3a] OK")
        else:
            result["phases_failed"].append(f"Phase3a:FundingSG — {err}")
            result["error"] = err
            print(f"  [Phase 3a] FAILED: {err}")
            return result

        # Phase 3b: Funding CIBC
        print("  [Phase 3b] Funding CIBC...")
        ok, err = _run_phase(
            "Phase3-CIBC",
            [sys.executable, "scripts/final_funding_cibc.py"],
            cwd=str(backend_dir),
            env=funding_env,
            timeout=300,
        )
        if ok:
            result["phases_ok"].append("Phase3b:FundingCIBC")
            print("  [Phase 3b] OK")
        else:
            result["phases_failed"].append(f"Phase3b:FundingCIBC — {err}")
            result["error"] = err
            print(f"  [Phase 3b] FAILED: {err}")
            return result

        # Phase 4: CashFlow (SG and CIBC)
        files_req = work_dir / "files_required"
        output_dir = work_dir / "output"
        curr_date = date_vars["curr_date"]

        for buyer in ("sg", "cibc"):
            sfy_file = files_req / f"FX3_{curr_date}_ExhibitAtoFormofSaleNotice_{buyer}.xlsx"
            prime_file = files_req / f"{curr_date} Exhibit A To Form Of Sale Notice_{buyer}.xlsx"

            if not sfy_file.exists() or not prime_file.exists():
                print(f"  [Phase 4 {buyer.upper()}] Skipped — exhibit files not found")
                continue

            print(f"  [Phase 4 {buyer.upper()}] CashFlow...")
            ok, err = _run_phase(
                f"Phase4-{buyer.upper()}",
                [
                    sys.executable, "-m", "cashflow.compute.run_purchase_package",
                    "--prime-file", str(prime_file),
                    "--sfy-file", str(sfy_file),
                    "--master-sheet", str(files_req / "MASTER_SHEET.xlsx"),
                    "--notes-sheet", str(files_req / "MASTER_SHEET - Notes.xlsx"),
                    "--purchase-date", pdate,
                    "--output-dir", str(output_dir),
                    "--buy-num", buy_num,
                    "--buyer", buyer,
                ],
                cwd=str(backend_dir),
                timeout=300,
            )
            if ok:
                result["phases_ok"].append(f"Phase4:CashFlow{buyer.upper()}")
                print(f"  [Phase 4 {buyer.upper()}] OK")
            else:
                result["phases_failed"].append(f"Phase4:CashFlow{buyer.upper()} — {err}")
                print(f"  [Phase 4 {buyer.upper()}] FAILED: {err}")

        # Diff outputs vs golden
        expected_outputs = test_case_dir / "output"
        if expected_outputs.exists():
            d, m, e = _diff_directories(output_dir, expected_outputs, "output")
            result["diffs"].extend(d)
            result["missing"].extend(m)
            result["extra"].extend(e)

        expected_share = test_case_dir / "output_share"
        if expected_share.exists():
            d, m, e = _diff_directories(work_dir / "output_share", expected_share, "output_share")
            result["diffs"].extend(d)
            result["missing"].extend(m)
            result["extra"].extend(e)

        total_issues = len(result["diffs"]) + len(result["missing"]) + len(result["extra"])
        result["status"] = "PASS" if total_issues == 0 else "FAILED"

        if total_issues == 0:
            print("  Diff result: PASS (0 diffs)")
        else:
            print(f"  Diff result: FAIL ({total_issues} differences)")
            for item in result["diffs"]:
                print(f"    DIFF: {item}")
            for item in result["missing"]:
                print(f"    MISSING: {item}")
            for item in result["extra"]:
                print(f"    EXTRA: {item}")

    except Exception as exc:
        result["error"] = str(exc)
        print(f"  FAILED (exception: {exc})")
    finally:
        if not no_cleanup and result.get("work_dir") is not None:
            try:
                shutil.rmtree(str(tmp_parent))
            except OSError as exc:
                print(f"  [WARN] Could not clean up {tmp_parent}: {exc}")

    return result


# ---------------------------------------------------------------------------
# Excel report generation
# ---------------------------------------------------------------------------

def _file_status(rel_path: str, result: dict) -> str:
    """Return status string for a file path like 'output/foo.xlsx'."""
    if rel_path in result["diffs"]:
        return "DIFFER"
    if rel_path in result["missing"]:
        return "MISSING"
    if rel_path in result["extra"]:
        return "EXTRA"
    return "MATCH"


def write_excel_report(results: list, report_path: Path) -> None:
    """Write test results to Excel following the TestMatrix template format."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Fills
    FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")   # MATCH / PASS
    FILL_RED    = PatternFill("solid", fgColor="FFC7CE")   # MISSING / FAILED
    FILL_ORANGE = PatternFill("solid", fgColor="FFCC99")   # DIFFER
    FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")   # EXTRA
    FILL_BLUE   = PatternFill("solid", fgColor="BDD7EE")   # header
    FILL_DKBLUE = PatternFill("solid", fgColor="4472C4")   # section header
    FILL_GRAY   = PatternFill("solid", fgColor="EEEEEE")   # baseline/input rows
    FILL_NONE   = PatternFill("none")

    FONT_BOLD  = Font(bold=True)
    FONT_WHITE = Font(bold=True, color="FFFFFF")
    ALIGN_CTR  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="AAAAAA")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _cell(ws, row, col, value=None, fill=None, font=None, align=None, border=None):
        c = ws.cell(row=row, column=col, value=value)
        if fill:   c.fill   = fill
        if font:   c.font   = font
        if align:  c.alignment = align
        if border: c.border = border
        return c

    def _status_fill(status: str) -> PatternFill:
        return {
            "MATCH":   FILL_GREEN,
            "DIFFER":  FILL_ORANGE,
            "MISSING": FILL_RED,
            "EXTRA":   FILL_YELLOW,
            "PASS":    FILL_GREEN,
            "FAILED":  FILL_RED,
        }.get(status, FILL_NONE)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # -----------------------------------------------------------------------
    # Summary sheet
    # -----------------------------------------------------------------------
    sum_ws = wb.create_sheet("Summary")
    sum_ws.column_dimensions["A"].width = 22
    for col_letter in ["B", "C", "D", "E", "F"]:
        sum_ws.column_dimensions[col_letter].width = 18

    # Title
    title_cell = sum_ws.cell(row=1, column=1, value="Test Summary")
    title_cell.font = Font(bold=True, size=14)
    sum_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(results))

    # Column headers (test case names)
    _cell(sum_ws, 2, 1, "Metric", fill=FILL_BLUE, font=FONT_BOLD, align=ALIGN_CTR, border=BORDER)
    for col_idx, r in enumerate(results, start=2):
        _cell(sum_ws, 2, col_idx, r["name"], fill=FILL_BLUE, font=FONT_BOLD, align=ALIGN_CTR, border=BORDER)

    rows_data = [
        ("pdate",          lambda r: r["pdate"]),
        ("tday",           lambda r: r["tday"]),
        ("Test Status",    lambda r: r["status"]),
        ("Phases OK",      lambda r: len(r["phases_ok"])),
        ("Phases Failed",  lambda r: len(r["phases_failed"])),
        ("Output Match",   lambda r: sum(1 for f in _all_expected_files(r) if _file_status(f, r) == "MATCH")),
        ("Output Differ",  lambda r: len(r["diffs"])),
        ("Output Missing", lambda r: len(r["missing"])),
        ("Output Extra",   lambda r: len(r["extra"])),
        ("Total Issues",   lambda r: len(r["diffs"]) + len(r["missing"]) + len(r["extra"])),
    ]

    for row_offset, (label, fn) in enumerate(rows_data, start=3):
        _cell(sum_ws, row_offset, 1, label, fill=FILL_GRAY, font=FONT_BOLD,
              align=ALIGN_LEFT, border=BORDER)
        for col_idx, r in enumerate(results, start=2):
            val = fn(r)
            fill = FILL_NONE
            if label == "Test Status":
                fill = _status_fill(str(val))
            elif label == "Total Issues":
                fill = FILL_GREEN if val == 0 else FILL_RED
            _cell(sum_ws, row_offset, col_idx, val, fill=fill, align=ALIGN_CTR, border=BORDER)

    # Phase breakdown rows
    sum_ws.cell(row=len(rows_data) + 4, column=1, value="Phases").font = Font(bold=True, italic=True)
    all_phases = []
    for r in results:
        for p in r["phases_ok"] + [p.split(" — ")[0] for p in r["phases_failed"]]:
            if p not in all_phases:
                all_phases.append(p)

    for row_offset, phase in enumerate(all_phases, start=len(rows_data) + 5):
        _cell(sum_ws, row_offset, 1, phase, fill=FILL_GRAY, align=ALIGN_LEFT, border=BORDER)
        for col_idx, r in enumerate(results, start=2):
            ok_names = [p.split(":")[0] + ":" + p.split(":")[1] if ":" in p else p for p in r["phases_ok"]]
            fail_names = [p.split(" — ")[0] for p in r["phases_failed"]]
            if phase in r["phases_ok"]:
                _cell(sum_ws, row_offset, col_idx, "OK", fill=FILL_GREEN, align=ALIGN_CTR, border=BORDER)
            elif phase in fail_names:
                _cell(sum_ws, row_offset, col_idx, "FAILED", fill=FILL_RED, align=ALIGN_CTR, border=BORDER)
            else:
                _cell(sum_ws, row_offset, col_idx, "—", fill=FILL_NONE, align=ALIGN_CTR, border=BORDER)

    # -----------------------------------------------------------------------
    # Per-test sheets
    # -----------------------------------------------------------------------
    for result in results:
        test_case_dir = result["test_case_dir"]
        sheet_name = result["name"][:31]  # Excel sheet name limit
        ws = wb.create_sheet(sheet_name)

        # Column widths
        ws.column_dimensions["A"].width = 12   # Item
        ws.column_dimensions["B"].width = 48   # File Name
        for i in range(len(PHASE_COLS)):
            ws.column_dimensions[get_column_letter(3 + i)].width = 18

        # Row 1: phase group headers
        _cell(ws, 1, 1, "Item",      fill=FILL_BLUE, font=FONT_BOLD, align=ALIGN_CTR, border=BORDER)
        _cell(ws, 1, 2, "File Name", fill=FILL_BLUE, font=FONT_BOLD, align=ALIGN_LEFT, border=BORDER)
        for i, phase_label in enumerate(PHASE_COLS):
            _cell(ws, 1, 3 + i, phase_label,
                  fill=FILL_BLUE, font=FONT_BOLD, align=ALIGN_CTR, border=BORDER)
        ws.row_dimensions[1].height = 36

        # Row 2: test metadata
        meta = (
            f"pdate={result['pdate']}  "
            f"tday={result['tday']}  "
            f"curr_date={result['date_vars']['curr_date']}  "
            f"buy_num={result['buy_num']}"
        )
        ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=2 + len(PHASE_COLS))
        _cell(ws, 2, 1, "Dates", fill=FILL_GRAY, font=FONT_BOLD, align=ALIGN_CTR)
        meta_cell = ws.cell(row=2, column=2, value=meta)
        meta_cell.font = Font(italic=True)
        meta_cell.alignment = ALIGN_LEFT

        current_row = 3

        # ---- INPUTS section ------------------------------------------------
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=2 + len(PHASE_COLS))
        sec_cell = ws.cell(row=current_row, column=1, value="INPUT FILES (files_required/)")
        sec_cell.fill = FILL_DKBLUE
        sec_cell.font = FONT_WHITE
        sec_cell.alignment = ALIGN_CTR
        current_row += 1

        # List all files in test_case_dir/files_required/
        files_required_dir = test_case_dir / "files_required"
        input_files = sorted(f.name for f in files_required_dir.iterdir() if f.is_file())

        for idx, fname in enumerate(input_files, start=1):
            phase_key = _classify_input_file(fname)
            phase_idx = PHASE_KEYS.index(phase_key) if phase_key in PHASE_KEYS else -1

            _cell(ws, current_row, 1, f"Input {idx}", fill=FILL_GRAY, align=ALIGN_CTR, border=BORDER)
            _cell(ws, current_row, 2, fname, fill=FILL_GRAY, align=ALIGN_LEFT, border=BORDER)
            for i in range(len(PHASE_COLS)):
                if i == phase_idx:
                    _cell(ws, current_row, 3 + i, "PRESENT",
                          fill=FILL_GREEN, align=ALIGN_CTR, border=BORDER)
                else:
                    _cell(ws, current_row, 3 + i, "",
                          fill=FILL_NONE, align=ALIGN_CTR, border=BORDER)
            current_row += 1

        # ---- EXPECTED OUTPUTS section --------------------------------------
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=2 + len(PHASE_COLS))
        sec_cell = ws.cell(row=current_row, column=1, value="EXPECTED OUTPUT FILES (output/ and output_share/)")
        sec_cell.fill = FILL_DKBLUE
        sec_cell.font = FONT_WHITE
        sec_cell.alignment = ALIGN_CTR
        current_row += 1

        # Collect all expected files from golden output/ and output_share/
        expected_output_files = []
        for subdir_label in ("output", "output_share"):
            golden_dir = test_case_dir / subdir_label
            if golden_dir.exists():
                for f in sorted(golden_dir.rglob("*")):
                    if f.is_file():
                        rel = f"{subdir_label}/{f.relative_to(golden_dir)}"
                        expected_output_files.append(rel)

        for idx, rel_path in enumerate(expected_output_files, start=1):
            fname = Path(rel_path).name
            status = _file_status(rel_path, result)
            phase_key = _classify_output_file(fname)
            phase_idx = PHASE_KEYS.index(phase_key) if phase_key in PHASE_KEYS else -1
            status_fill = _status_fill(status)
            row_fill = FILL_GRAY if status == "MATCH" else FILL_NONE

            _cell(ws, current_row, 1, f"Output {idx}", fill=row_fill, align=ALIGN_CTR, border=BORDER)
            _cell(ws, current_row, 2, rel_path, fill=row_fill, align=ALIGN_LEFT, border=BORDER)
            for i in range(len(PHASE_COLS)):
                if i == phase_idx:
                    _cell(ws, current_row, 3 + i, status,
                          fill=status_fill, align=ALIGN_CTR, border=BORDER)
                else:
                    _cell(ws, current_row, 3 + i, "",
                          fill=FILL_NONE, align=ALIGN_CTR, border=BORDER)
            current_row += 1

        # ---- EXTRA FILES section (generated but not expected) --------------
        extra_files = [e for e in result["extra"]]
        if extra_files:
            current_row += 1
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=2 + len(PHASE_COLS))
            sec_cell = ws.cell(row=current_row, column=1,
                               value="EXTRA FILES (generated but not in golden)")
            sec_cell.fill = FILL_DKBLUE
            sec_cell.font = FONT_WHITE
            sec_cell.alignment = ALIGN_CTR
            current_row += 1

            for idx, rel_path in enumerate(extra_files, start=1):
                fname = Path(rel_path).name
                phase_key = _classify_output_file(fname)
                phase_idx = PHASE_KEYS.index(phase_key) if phase_key in PHASE_KEYS else -1

                _cell(ws, current_row, 1, f"Extra {idx}", fill=FILL_YELLOW,
                      align=ALIGN_CTR, border=BORDER)
                _cell(ws, current_row, 2, rel_path, fill=FILL_YELLOW,
                      align=ALIGN_LEFT, border=BORDER)
                for i in range(len(PHASE_COLS)):
                    if i == phase_idx:
                        _cell(ws, current_row, 3 + i, "EXTRA",
                              fill=FILL_YELLOW, align=ALIGN_CTR, border=BORDER)
                    else:
                        _cell(ws, current_row, 3 + i, "",
                              fill=FILL_NONE, align=ALIGN_CTR, border=BORDER)
                current_row += 1

        # ---- PHASE STATUS section ------------------------------------------
        current_row += 1
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=2 + len(PHASE_COLS))
        sec_cell = ws.cell(row=current_row, column=1, value="PHASE EXECUTION STATUS")
        sec_cell.fill = FILL_DKBLUE
        sec_cell.font = FONT_WHITE
        sec_cell.alignment = ALIGN_CTR
        current_row += 1

        phase_ok_set = set(result["phases_ok"])
        fail_map = {p.split(" — ")[0]: p.split(" — ", 1)[1] if " — " in p else ""
                    for p in result["phases_failed"]}

        phase_display = [
            ("Phase 1: Pre-Funding",  "Phase1:PreFunding"),
            ("Phase 2: Tagging",      "Phase2:Tagging"),
            ("Phase 3a: Funding SG",  "Phase3a:FundingSG"),
            ("Phase 3b: Funding CIBC","Phase3b:FundingCIBC"),
            ("Phase 4: CashFlow SG",  "Phase4:CashFlowSG"),
            ("Phase 4: CashFlow CIBC","Phase4:CashFlowCIBC"),
        ]
        for phase_label, phase_key in phase_display:
            if phase_key in phase_ok_set or f"{phase_key}(skipped)" in phase_ok_set:
                status_val = "OK (skipped)" if f"{phase_key}(skipped)" in phase_ok_set else "OK"
                s_fill = FILL_GREEN
            elif phase_key in fail_map:
                status_val = f"FAILED: {fail_map[phase_key]}"
                s_fill = FILL_RED
            else:
                status_val = "—"
                s_fill = FILL_NONE

            _cell(ws, current_row, 1, phase_label, fill=FILL_GRAY,
                  font=FONT_BOLD, align=ALIGN_LEFT, border=BORDER)
            ws.merge_cells(start_row=current_row, start_column=2,
                           end_row=current_row, end_column=2 + len(PHASE_COLS))
            _cell(ws, current_row, 2, status_val, fill=s_fill, align=ALIGN_LEFT, border=BORDER)
            current_row += 1

        # ---- SUMMARY row ---------------------------------------------------
        current_row += 1
        total_issues = len(result["diffs"]) + len(result["missing"]) + len(result["extra"])
        overall_status = "PASS" if result["status"] == "PASS" else f"FAILED ({total_issues} issues)"
        _cell(ws, current_row, 1, "Overall", fill=FILL_GRAY, font=FONT_BOLD, align=ALIGN_CTR, border=BORDER)
        ws.merge_cells(start_row=current_row, start_column=2,
                       end_row=current_row, end_column=2 + len(PHASE_COLS))
        _cell(ws, current_row, 2, overall_status,
              fill=_status_fill(result["status"]), font=FONT_BOLD, align=ALIGN_CTR, border=BORDER)

    wb.save(str(report_path))
    print(f"\nExcel report written: {report_path}")


def _all_expected_files(result: dict) -> list[str]:
    """Return all expected output file paths (relative, prefixed with output/ or output_share/)."""
    files = []
    test_case_dir = result.get("test_case_dir")
    if test_case_dir is None:
        return files
    for subdir_label in ("output", "output_share"):
        golden_dir = Path(test_case_dir) / subdir_label
        if golden_dir.exists():
            for f in golden_dir.rglob("*"):
                if f.is_file():
                    rel = f"{subdir_label}/{f.relative_to(golden_dir)}"
                    files.append(rel)
    return files


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Data regression harness: runs all four pipeline phases against each "
            "buy-date folder in TestData and diffs outputs against expected values. "
            "Generates an Excel report following the TestMatrix template format."
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=r"""
Phases run per test case:
  1. Pre-Funding  (run_pipeline_cli.py)
  2. Tagging      (tagging.py)
  3. Funding SG   (final_funding_sg.py)
  4. Funding CIBC (final_funding_cibc.py)
  5. CashFlow     (cashflow.compute.run_purchase_package for SG and CIBC)

Examples:
  python backend/scripts/regression_test.py
  python backend/scripts/regression_test.py --test-data C:\Users\omack\Downloads\TestData
  python backend/scripts/regression_test.py --pdate 2026-02-24 --tday 2026-02-19
  python backend/scripts/regression_test.py --no-cleanup
  python backend/scripts/regression_test.py --skip-phase1
  python backend/scripts/regression_test.py --report C:\temp\results.xlsx
""",
    )
    parser.add_argument(
        "--test-data",
        type=str,
        default=r"C:\Users\omack\Downloads\TestData",
        help=r"Root directory containing buy-date test case folders.",
    )
    parser.add_argument(
        "--backend-dir",
        type=str,
        default=None,
        help="Path to the backend/ directory. Default: auto-detected.",
    )
    parser.add_argument(
        "--pdate",
        type=str,
        default=None,
        help="Purchase date YYYY-MM-DD. Applied to ALL test cases.",
    )
    parser.add_argument(
        "--tday",
        type=str,
        default=None,
        help="Base date YYYY-MM-DD for file naming. Applied to ALL test cases.",
    )
    parser.add_argument(
        "--no-cleanup",
        action="store_true",
        help="Do not delete temp work directories after each test case.",
    )
    parser.add_argument(
        "--skip-phase1",
        action="store_true",
        help="Skip Phase 1 (pre-funding). Useful when files_required/ already contains "
             "pre-generated exhibit files.",
    )
    parser.add_argument(
        "--report",
        type=str,
        default=None,
        help="Path for the Excel report output. Default: TestMatrix_Results_<timestamp>.xlsx "
             "written to the test-data directory.",
    )
    args = parser.parse_args()

    test_data_dir = Path(args.test_data)
    backend_dir = Path(args.backend_dir) if args.backend_dir else _default_backend_dir()

    if not backend_dir.exists():
        print(f"ERROR: backend directory not found: {backend_dir}", file=sys.stderr)
        sys.exit(1)

    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    print("=" * 60)
    print(f"REGRESSION TEST REPORT — {now_str}")
    print(f"Test data: {test_data_dir}")
    print("=" * 60)

    test_cases = discover_test_cases(test_data_dir)
    dates_config = _load_dates_config(test_data_dir)

    if not test_cases:
        print("\nNo valid test cases found. Nothing to run.")
        print("\n" + "=" * 60)
        print("SUMMARY: 0 PASSED / 0 FAILED / 0 TOTAL")
        print("=" * 60)
        sys.exit(0)

    results = []
    for case_dir in test_cases:
        result = run_test_case(
            test_case_dir=case_dir,
            backend_dir=backend_dir,
            cli_pdate=args.pdate,
            cli_tday=args.tday,
            no_cleanup=args.no_cleanup,
            skip_phase1=args.skip_phase1,
            dates_config=dates_config,
        )
        results.append(result)

    # Final summary
    passed = sum(1 for r in results if r["status"] == "PASS")
    failed = sum(1 for r in results if r["status"] != "PASS")
    total = len(results)

    print("\n" + "=" * 60)
    print(f"SUMMARY: {passed} PASSED / {failed} FAILED / {total} TOTAL")
    print("=" * 60)

    if failed > 0:
        print("\nFailed cases:")
        for r in results:
            if r["status"] != "PASS":
                err_note = f" — {r['error']}" if r["error"] else ""
                phases_note = (
                    f" [failed phases: {', '.join(r['phases_failed'])}]"
                    if r["phases_failed"] else ""
                )
                print(f"  FAIL: {r['name']}{err_note}{phases_note}")

    # Write Excel report
    if args.report:
        report_path = Path(args.report)
    else:
        report_path = test_data_dir / f"TestMatrix_Results_{timestamp}.xlsx"

    try:
        write_excel_report(results, report_path)
    except Exception as exc:
        print(f"[WARN] Could not write Excel report: {exc}", file=sys.stderr)

    sys.exit(0 if failed == 0 else 1)


if __name__ == "__main__":
    main()
