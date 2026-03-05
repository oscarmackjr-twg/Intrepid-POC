"""
Generate minimal synthetic sample data for local dev onboarding.

Usage:
    python scripts/create_sample_data.py

Outputs to: backend/data/sample/files_required/

File naming convention:
- pdate = 2026-02-19
- yesterday = 2026-02-18  →  filename date = 02-18-2026
"""
import csv
import os
from pathlib import Path

import openpyxl
from openpyxl import Workbook


OUTPUT_DIR = Path(__file__).parent.parent / "data" / "sample" / "files_required"


def create_master_sheet():
    """MASTER_SHEET.xlsx — loan program lookup table used by enrich_buy_df().
    Pipeline merges buy_df on ['loan program', 'Platform'] columns.
    Columns: loan program, platform, type, new_programs, ...
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "loan program",
        "platform",
        "type",
        "new_programs",
        "Dealer Fee",
        "Description",
    ]
    ws.append(headers)

    rows = [
        ["FX3_PRIME_STD", "PRIME", "standard", False, 1.5, "Prime standard loan"],
        ["FX3_SFY_STD", "SFY", "standard", False, 1.5, "SFY standard loan"],
        ["FX3_PRIME_HD", "PRIME", "hd_note", False, 2.0, "Prime HD note"],
        ["FX3_SFY_HD", "SFY", "hd_note", False, 2.0, "SFY HD note"],
        ["FX3_PRIME_NEW", "PRIME", "standard", True, 1.5, "Prime new program"],
    ]
    for row in rows:
        ws.append(row)

    path = OUTPUT_DIR / "MASTER_SHEET.xlsx"
    wb.save(path)
    print(f"Created: {path}")


def create_master_sheet_notes():
    """MASTER_SHEET - Notes.xlsx — notes variant of loan program lookup.
    Same structure as MASTER_SHEET but for notes loans.
    Pipeline appends 'notes' suffix to loan program before merging.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    headers = [
        "loan program",
        "platform",
        "type",
        "new_programs",
        "Dealer Fee",
        "Description",
    ]
    ws.append(headers)

    rows = [
        ["FX3_PRIME_STDnotes", "PRIME", "standard", False, 1.5, "Prime standard note"],
        ["FX3_SFY_STDnotes", "SFY", "standard", False, 1.5, "SFY standard note"],
        ["FX3_PRIME_HDnotes", "PRIME", "hd_note", False, 2.0, "Prime HD note variant"],
    ]
    for row in rows:
        ws.append(row)

    path = OUTPUT_DIR / "MASTER_SHEET - Notes.xlsx"
    wb.save(path)
    print(f"Created: {path}")


def create_current_assets():
    """current_assets.csv — existing portfolio assets used by pipeline.
    Pipeline does:
      existing_file['Submit Date'] = pd.to_datetime(existing_file['Submit Date'])
      existing_file['Purchase_Date'] = pd.to_datetime(existing_file['Purchase_Date'])
      existing_file['Monthly Payment Date'] = pd.to_datetime(existing_file['Monthly Payment Date'])
      ... uses 'SELLER Loan #', 'Repurchase', 'Purchase_Date', 'Orig. Balance', 'platform'
    Eligibility checks filter: existing_file['Purchase_Date'] > '2025-10-01'
    """
    headers = [
        "SELLER Loan #",
        "Account Number",
        "Submit Date",
        "Purchase_Date",
        "Monthly Payment Date",
        "Orig. Balance",
        "platform",
        "loan program",
        "Application Type",
        "Lender Price(%)",
        "FICO Borrower",
        "DTI",
        "PTI",
        "Term",
        "APR",
        "Property State",
        "Repurchase",
        "type",
        "purchase_price_check",
    ]

    rows = [
        ["SFC_1001", 1001, "2025-11-01", "2025-11-05", "2025-12-01",
         150000.00, "prime", "FX3_PRIME_STD", "Standard", 98.50,
         720, 35.0, 20.0, 120, 6.25, "CA", False, "standard", True],
        ["SFC_1002", 1002, "2025-11-15", "2025-11-20", "2025-12-15",
         200000.00, "sfy", "FX3_SFY_STD", "Standard", 97.75,
         680, 40.0, 22.0, 180, 6.75, "TX", False, "standard", True],
        ["SFC_1003", 1003, "2025-12-01", "2025-12-10", "2026-01-01",
         175000.00, "prime", "FX3_PRIME_STD", "Standard", 98.00,
         700, 38.0, 21.0, 120, 6.50, "FL", False, "standard", False],
    ]

    path = OUTPUT_DIR / "current_assets.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"Created: {path}")


def create_underwriting_grids():
    """Underwriting_Grids_COMAP.xlsx — multi-sheet workbook.
    Required sheets (from pipeline.py load_reference_data):
      SFY, Prime, SFY - Notes, Prime - Notes,
      SFY COMAP, SFY COMAP2, Prime COMAP, Notes CoMAP,
      SFY COMAP-Oct25, SFY COMAP-Oct25-2,
      Prime CoMAP-Oct25, Prime CoMAP-Oct25-2, Prime CoMAP - New
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Underwriting sheets: SFY, Prime, SFY - Notes, Prime - Notes
    uw_headers = ["Term", "FICO", "DTI", "PTI", "Max Balance", "Note"]
    uw_sheets = ["SFY", "Prime", "SFY - Notes", "Prime - Notes"]
    for sheet_name in uw_sheets:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(uw_headers)
        ws.append([120, 700, 40.0, 22.0, 300000, "Standard guideline"])
        ws.append([180, 680, 45.0, 25.0, 400000, "Extended term guideline"])

    # CoMAP sheets: SFY COMAP, SFY COMAP2, Prime COMAP, Notes CoMAP
    comap_headers = ["loan program", "Platform", "Eligible", "Max LTV", "Min FICO", "Note"]
    comap_sheets = [
        "SFY COMAP", "SFY COMAP2", "Prime COMAP", "Notes CoMAP",
        "SFY COMAP-Oct25", "SFY COMAP-Oct25-2",
        "Prime CoMAP-Oct25", "Prime CoMAP-Oct25-2", "Prime CoMAP - New",
    ]
    for sheet_name in comap_sheets:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(comap_headers)
        ws.append(["FX3_PRIME_STD", "PRIME", True, 95.0, 680, "Standard prime"])
        ws.append(["FX3_SFY_STD", "SFY", True, 95.0, 680, "Standard SFY"])
        ws.append(["FX3_PRIME_HD", "PRIME", True, 90.0, 700, "Prime HD note"])
        ws.append(["FX3_SFY_HD", "SFY", True, 90.0, 700, "SFY HD note"])

    path = OUTPUT_DIR / "Underwriting_Grids_COMAP.xlsx"
    wb.save(path)
    print(f"Created: {path}")


def create_tape_loans():
    """Tape20Loans_02-18-2026.csv — loan tape input file.
    pipeline.py: normalize_loans_df() requires:
      'Account Number', 'Loan Group', 'Status Codes'
    enrichment.py adds SELLER Loan # from Account Number
    tag_loans_by_group: tags SFY if 'FX3' or 'FX1' in Loan Group
    pipeline.py: existing_file['SELLER Loan #'].isin(repurchased_loans)
    """
    headers = [
        "Account Number",
        "Loan Group",
        "Status Codes",
        "Open Date",
        "maturityDate",
        "Submit Date",
        "Purchase_Date",
        "Monthly Payment Date",
        "Orig. Balance",
        "platform",
        "loan program",
        "Application Type",
        "Lender Price(%)",
        "Purchase Price",
        "FICO Borrower",
        "DTI",
        "PTI",
        "Term",
        "APR",
        "Property State",
    ]

    rows = [
        [2001, "FX3_PRIME", "", "2026-01-15", "2028-01-15",
         "2026-02-01", "2026-02-19", "2026-03-01",
         175000.00, "prime", "FX3_PRIME_STD", "Standard",
         98.50, 172375.00, 720, 35.0, 20.0, 120, 6.25, "CA"],
        [2002, "FX3_SFY", "", "2026-01-20", "2030-01-20",
         "2026-02-05", "2026-02-19", "2026-03-05",
         225000.00, "sfy", "FX3_SFY_STD", "Standard",
         97.75, 219937.50, 680, 40.0, 22.0, 180, 6.75, "TX"],
        [2003, "FX3_PRIME", "", "2026-01-25", "2028-01-25",
         "2026-02-10", "2026-02-19", "2026-03-10",
         195000.00, "prime", "FX3_PRIME_STD", "Standard",
         98.00, 191100.00, 700, 38.0, 21.0, 120, 6.50, "FL"],
        [2004, "FX1_SFY", "REPURCHASE", "2025-09-01", "2027-09-01",
         "2025-09-15", "2025-10-01", "2025-11-01",
         155000.00, "sfy", "FX3_SFY_STD", "HD NOTE",
         97.00, 150350.00, 660, 42.0, 24.0, 120, 7.00, "NY"],
        [2005, "FX3_PRIME", "", "2026-02-01", "2029-02-01",
         "2026-02-12", "2026-02-19", "2026-03-12",
         310000.00, "prime", "FX3_PRIME_NEW", "Standard",
         98.25, 304575.00, 750, 32.0, 18.0, 120, 6.10, "WA"],
    ]

    path = OUTPUT_DIR / "Tape20Loans_02-18-2026.csv"
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"Created: {path}")


def create_sfy_exhibit():
    """SFY_02-18-2026_ExhibitAtoFormofSaleNotice - Pre-Funding.xlsx
    normalize_sfy_df(): skips first 4 rows, then uses row 5 as header.
    So we need at least 6 rows: 4 filler rows, 1 header row, 1+ data rows.
    Columns used by pipeline: SELLER Loan #, Loan Program, Application Type,
    Platform (added later), TU144 (optional).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 4 filler rows (skipped by normalize_sfy_df)
    for i in range(4):
        ws.append([f"Filler row {i+1}", "", "", "", "", ""])

    # Row 5 becomes the header after iloc[4:] and columns = df.iloc[0]
    ws.append([
        "SELLER Loan #", "Loan Program", "Application Type",
        "Orig. Balance", "Lender Price(%)", "Purchase Price",
        "FICO Borrower", "DTI", "PTI", "Term", "APR", "Property State",
        "Submit Date", "TU144",
    ])

    # Data rows (row 6 onward)
    ws.append([
        "SFC_2002", "FX3_SFY_STD", "Standard",
        225000.00, 97.75, 219937.50,
        680, 40.0, 22.0, 180, 6.75, "TX",
        "2026-02-05", 0,
    ])
    ws.append([
        "SFC_2004", "FX3_SFY_STD", "HD NOTE",
        155000.00, 97.00, 150350.00,
        660, 42.0, 24.0, 120, 7.00, "NY",
        "2025-09-15", 0,
    ])

    path = OUTPUT_DIR / "SFY_02-18-2026_ExhibitAtoFormofSaleNotice - Pre-Funding.xlsx"
    wb.save(path)
    print(f"Created: {path}")


def create_prime_exhibit():
    """PRIME_02-18-2026_ExhibitAtoFormofSaleNotice - Pre-Funding.xlsx
    normalize_prime_df(): same 4-row skip pattern as SFY.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 4 filler rows (skipped by normalize_prime_df)
    for i in range(4):
        ws.append([f"Filler row {i+1}", "", "", "", "", ""])

    # Row 5 becomes the header
    ws.append([
        "SELLER Loan #", "Loan Program", "Application Type",
        "Orig. Balance", "Lender Price(%)", "Purchase Price",
        "FICO Borrower", "DTI", "PTI", "Term", "APR", "Property State",
        "Submit Date",
    ])

    # Data rows
    ws.append([
        "SFC_2001", "FX3_PRIME_STD", "Standard",
        175000.00, 98.50, 172375.00,
        720, 35.0, 20.0, 120, 6.25, "CA",
        "2026-02-01",
    ])
    ws.append([
        "SFC_2003", "FX3_PRIME_STD", "Standard",
        195000.00, 98.00, 191100.00,
        700, 38.0, 21.0, 120, 6.50, "FL",
        "2026-02-10",
    ])
    ws.append([
        "SFC_2005", "FX3_PRIME_NEW", "Standard",
        310000.00, 98.25, 304575.00,
        750, 32.0, 18.0, 120, 6.10, "WA",
        "2026-02-12",
    ])

    path = OUTPUT_DIR / "PRIME_02-18-2026_ExhibitAtoFormofSaleNotice - Pre-Funding.xlsx"
    wb.save(path)
    print(f"Created: {path}")


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Writing sample data to: {OUTPUT_DIR.resolve()}")

    create_master_sheet()
    create_master_sheet_notes()
    create_current_assets()
    create_underwriting_grids()
    create_tape_loans()
    create_sfy_exhibit()
    create_prime_exhibit()

    print(f"\nDone. Files in {OUTPUT_DIR}:")
    for f in sorted(OUTPUT_DIR.iterdir()):
        print(f"  {f.name}")


if __name__ == "__main__":
    main()
