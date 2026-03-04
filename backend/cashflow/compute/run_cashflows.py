"""Cash flow pipeline: current_assets.csv → multi-sheet Excel workbook.

Usage
-----
    python -m compute.cashflow.run_cashflows --input <path/to/current_assets.csv>

Options
-------
    --input   Path to current_assets.csv (required).
    --output  Path for the output .xlsx file.
              Default: same directory as --input, named cashflows_<date>.xlsx
    --cprshock  CPR multiplier shock (default 1.0 = no shock).
    --cdrshock  CDR multiplier shock (default 1.0 = no shock).
    --target  IRR target in percent for annual IRR-support solver (default 7.9).

What it produces
----------------
    Sheet "BD Cashflows"       — date-aggregated modeled cashflows, BD loans
    Sheet "Non-BD Cashflows"   — date-aggregated modeled cashflows, Non-BD loans
    Sheet "18 mth Stack"       — forward 18-month UPB composition by segment
    Sheet "MV UPB 18 mth Stack"— same, market-value weighted
    Sheet "PRIME Data"         — loan-level summary, PRIME platform
    Sheet "SFY Data"           — loan-level summary, SFY platform
    Sheet "IRR Support"        — cumulative annual IRR-support adjustments
"""
from __future__ import annotations

import argparse
import os
import sys
import tempfile
from concurrent.futures import ProcessPoolExecutor
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Callable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from pandas.tseries.offsets import DateOffset

from cashflow.compute.behavioral_model import (
    create_contractual_flow,
    price_loan_sfc_case,
    price_loan_our_case,
    xirr,
)

_ROW_FIELDS = [
    "SELLER Loan #",
    "Orig. Balance",
    "type",
    "promo_term",
    "Term",
    "APR",
    "coupon",
    "Monthly Payment",
    "Monthly Payment Date",
    "Property State",
    "modeled_purchase_price",
    "loan program",
    "platform",
    "tagging",
    "cpr",
    "constant_cpr",
    "cdr",
    "cdr_promo",
    "late_percent",
    "recovery_percent",
    "servicing_cost",
    "proposed_promo_cpr_percent",
    "proposed_cpr",
    "proposed_cdr",
    "proposed_cdr_promo",
]

_DATE_COLUMNS = ["Submit Date", "Purchase_Date", "Monthly Payment Date"]
_LOAN_SUMMARY_COLS = [
    "loan_id",
    "loan_program",
    "platform",
    "tagging",
    "loan_type",
    "Original Loan Amount",
    "modeled_purchase_price",
    "purchase_price",
    "IRR",
    "WAL",
    "SFC_IRR",
    "SFC_WAL",
    "SFC_gross_dlq",
    "SFC_gross_dlq_net_recovery",
    "OUR_gross_dlq",
    "OUR_gross_dlq_net_recovery",
]

# ---------------------------------------------------------------------------
# Input preparation helpers
# ---------------------------------------------------------------------------

def _resolve_loan_term(row: pd.Series) -> int:
    """Compute effective loan_term from loan type and raw Term column."""
    t = str(row["type"])
    if "ninp" in t:
        return int(row["Term"]) + int(row["promo_term"])
    if t == "epni":
        return int(row["promo_term"])
    return int(row["Term"])


def _resolve_coupon(row: pd.Series) -> float:
    """Resolve the coupon rate used for the amortizing schedule."""
    t = str(row["type"])
    if t in ("hybrid", "ninp"):
        coupon = float(row["APR"]) / 100.0 if t == "ninp" else float(row["coupon"]) / 10_000.0
    elif t == "epni":
        coupon = float(row["coupon"]) / 10_000.0
    else:
        coupon = float(row["APR"]) / 100.0
    # Connecticut usury cap: if coupon stored as basis-points*100 > 1200bp → cap at 11.99%
    if str(row.get("Property State", "")) == "CT" and float(row.get("coupon", 0)) > 1200:
        coupon = 0.1199
    return coupon


def _resolve_payment(row: pd.Series, loan_term: int) -> float:
    """Resolve monthly payment (epni uses principal / term)."""
    if str(row["type"]) == "epni":
        return float(row["Orig. Balance"]) / loan_term
    return float(row["Monthly Payment"])


def _resolve_loan_date(row: pd.Series) -> pd.Timestamp:
    """First actual payment date, adjusted back by promo months for ninp/hybrid."""
    pay_date = pd.to_datetime(row["Monthly Payment Date"])
    t = str(row["type"])
    if t in ("hybrid", "ninp"):
        return pay_date - DateOffset(months=int(row["promo_term"]))
    return pay_date


def _count_csv_rows(path: Path) -> int:
    """Count data rows in a CSV file without materializing it."""
    with path.open("r", encoding="utf-8", newline="") as handle:
        return max(sum(1 for _ in handle) - 1, 0)


# ---------------------------------------------------------------------------
# Per-loan processing
# ---------------------------------------------------------------------------

def _process_loan(row: pd.Series, cprshock: float, cdrshock: float):
    """Run both SFC and OUR case for a single loan row.

    Returns
    -------
    tuple: (sfc_df, our_df, loan_summary_dict)
    """
    loan_id = row["SELLER Loan #"]
    loan_amt = float(row["Orig. Balance"])
    loan_type = str(row["type"])
    promo_term = int(row["promo_term"])

    loan_term = _resolve_loan_term(row)
    coupon = _resolve_coupon(row)
    mth_payment = _resolve_payment(row, loan_term)
    loan_date = _resolve_loan_date(row)

    # Build contractual schedule
    contractual = create_contractual_flow(
        monthly_payment=mth_payment,
        loan_rate=coupon,
        loan_original_amount=loan_amt,
        loan_term=loan_term,
        loan_type=loan_type,
        promo_loan_term=promo_term,
        loan_id=loan_id,
        loan_date=loan_date,
    )

    # SFC (seller) case
    sfc_df, sfc_irr, sfc_wal = price_loan_sfc_case(
        row, contractual, cpr_shock=cprshock, cdr_shock=cdrshock,
    )

    # OUR (stressed) case
    our_df, our_irr, our_wal = price_loan_our_case(
        row, contractual, cpr_shock=cprshock, cdr_shock=cdrshock,
    )

    # Market-value UPB
    pp = float(row["modeled_purchase_price"])
    our_df["opening_mv_upb"] = our_df["opening_upb"] * pp
    sfc_df["opening_mv_upb"] = sfc_df["opening_upb"] * pp

    # Loan-level summary
    summary = {
        "loan_id": loan_id,
        "loan_program": row["loan program"],
        "platform": row["platform"],
        "tagging": row.get("tagging", ""),
        "loan_type": loan_type,
        "Original Loan Amount": loan_amt,
        "modeled_purchase_price": pp,
        "purchase_price": pp * loan_amt,
        "IRR": our_irr,
        "WAL": our_wal,
        "SFC_IRR": sfc_irr,
        "SFC_WAL": sfc_wal,
        "SFC_gross_dlq": sfc_df["write_off"].sum() / loan_amt,
        "SFC_gross_dlq_net_recovery": (
            (sfc_df["write_off"].sum() - sfc_df["recovery"].sum()) / loan_amt
        ),
        "OUR_gross_dlq": our_df["write_off"].sum() / loan_amt,
        "OUR_gross_dlq_net_recovery": (
            (our_df["write_off"].sum() - our_df["recovery"].sum()) / loan_amt
        ),
    }

    return sfc_df, our_df, summary


def _process_loan_task(payload):
    row, cprshock, cdrshock = payload
    try:
        sfc_df, our_df, summary = _process_loan(row, cprshock, cdrshock)
        tagging = row.get("tagging", "")
        sfc_df["tagging"] = tagging
        our_df["tagging"] = tagging
        return sfc_df, our_df, summary, None
    except Exception as exc:
        return None, None, None, f"{row['SELLER Loan #']} skipped — {exc}"


# ---------------------------------------------------------------------------
# Portfolio aggregation
# ---------------------------------------------------------------------------

_CASHFLOW_COLS = [
    "modeled_interest", "modeled_principal", "pre_payment", "write_off",
    "late_fee", "recovery", "servicing_cost", "total_principal_collected",
    "cash_adjustment", "wal_adjustment", "opening_upb", "opening_mv_upb",
]

_EMPTY_SEGMENT_COLS = ["dates"] + [f"__PENDING__{c}" for c in _CASHFLOW_COLS + ["int_minus_ca", "total_inflow"]]


def _normalize_excel_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().replace(tzinfo=None)
    return value


def _write_frame_sheet(workbook: Workbook, sheet_name: str, frame: pd.DataFrame) -> None:
    sheet = workbook.create_sheet(title=sheet_name)
    sheet.append(list(frame.columns))
    _append_frame_rows(sheet, frame)


def _append_frame_rows(sheet, frame: pd.DataFrame) -> None:
    for row in frame.itertuples(index=False, name=None):
        sheet.append([_normalize_excel_value(value) for value in row])


def _monthly_rollup(df: pd.DataFrame) -> pd.DataFrame:
    """Collapse one loan cashflow table to monthly totals."""
    loan_dates = pd.to_datetime(df["loan_dates"]).dt.to_period("M").dt.to_timestamp()
    return (
        df.groupby(loan_dates, sort=False, observed=True)[_CASHFLOW_COLS]
        .sum()
        .reset_index()
        .rename(columns={"loan_dates": "dates"})
    )


def _accumulate_rollup(accumulator: defaultdict, monthly_df: pd.DataFrame) -> None:
    """Add monthly cashflow rows into a dict-backed accumulator."""
    if monthly_df.empty:
        return
    values = monthly_df[_CASHFLOW_COLS].to_numpy(dtype=float, copy=False)
    for dt_value, row_values in zip(monthly_df["dates"].to_numpy(), values):
        accumulator[pd.Timestamp(dt_value)] += row_values


def _segment_frame_from_accumulator(accumulator: defaultdict, prefix: str) -> pd.DataFrame:
    """Convert a dict-backed accumulator into the workbook segment shape."""
    if not accumulator:
        return pd.DataFrame(columns=["dates"] + [
            f"{prefix}_{c}" for c in _CASHFLOW_COLS + ["int_minus_ca", "total_inflow"]
        ])

    dates = sorted(accumulator)
    data = np.vstack([accumulator[d] for d in dates])
    agg = pd.DataFrame(data, columns=_CASHFLOW_COLS)
    agg.insert(0, "dates", dates)
    agg["int_minus_ca"] = agg["modeled_interest"] - agg["cash_adjustment"]
    agg["total_inflow"] = (
        agg["modeled_interest"] + agg["modeled_principal"]
        + agg["pre_payment"] + agg["recovery"]
        - agg["cash_adjustment"] - agg["servicing_cost"]
    )
    agg.rename(columns={c: f"{prefix}_{c}" for c in _CASHFLOW_COLS + ["int_minus_ca", "total_inflow"]},
               inplace=True)
    return agg


# ---------------------------------------------------------------------------
# 18-month UPB forward stacks
# ---------------------------------------------------------------------------

def _build_upb_stack(
    df_bd: pd.DataFrame,
    df_non_bd: pd.DataFrame,
    upb_col_bd: str = "BD_opening_upb",
    upb_col_non_bd: str = "Non_BD_opening_upb",
    label_bd: str = "Pool BD UPB",
    label_non_bd: str = "Pool Non BD UPB",
) -> pd.DataFrame:
    """Create 18-month forward UPB stack.

    For each date t, the pool UPB is the sum of UPB contributions from
    loans funded over the preceding 18 months (forward-looking window).
    """
    merged = (
        pd.merge(
            df_bd[["dates", upb_col_bd]],
            df_non_bd[["dates", upb_col_non_bd]],
            on="dates",
            how="outer",
        )
        .fillna(0.0)
        .sort_values("dates")
    )
    if merged.empty:
        return pd.DataFrame(columns=["dates", label_bd, label_non_bd])

    merged = (
        merged.set_index("dates")
        .asfreq("MS", fill_value=0.0)
        .reset_index()
    )

    bd_vals = merged[upb_col_bd].to_numpy(dtype=float, copy=False)
    non_bd_vals = merged[upb_col_non_bd].to_numpy(dtype=float, copy=False)
    window = np.ones(18, dtype=float)
    bd_stack = np.convolve(bd_vals[::-1], window, mode="full")[: len(bd_vals)][::-1]
    non_bd_stack = np.convolve(non_bd_vals[::-1], window, mode="full")[: len(non_bd_vals)][::-1]

    return pd.DataFrame({
        "dates": merged["dates"].to_numpy(),
        label_bd: bd_stack,
        label_non_bd: non_bd_stack,
    })


# ---------------------------------------------------------------------------
# IRR support solver (annual)
# ---------------------------------------------------------------------------

def _solve_irr_support(
    irr_monthly_agg: pd.DataFrame,
    total_purchase: float,
    irr_target_pct: float,
) -> pd.DataFrame:
    """Binary search for annual reserve-cost addition to hit target IRR.

    For each year 1-10, finds the cumulative servicing-cost adjustment needed
    so that the portfolio XIRR meets irr_target_pct.

    Returns a DataFrame with columns: year, reserve_cost_cumsum,
    Reverse_Servicing_cost_addition.
    """
    agg = irr_monthly_agg.sort_values("dates").rename(columns={"dates": "loan_dates"}).reset_index(drop=True)
    if agg.empty:
        return pd.DataFrame(
            {"year": range(1, 11), "reserve_cost_cumsum": [0.0] * 10,
             "Reverse_Servicing_cost_addition": [0.0] * 10}
        )
    reserve_cumsum = 0.0
    rows = []

    for yr in range(1, 11):
        cutoff = agg["loan_dates"].iloc[0] + DateOffset(years=yr)
        window = agg[agg["loan_dates"] <= cutoff].copy()

        net_cf = (
            window["modeled_interest"] + window["modeled_principal"]
            + window["pre_payment"] + window["recovery"]
            - window["cash_adjustment"] - window["servicing_cost"]
        ).values
        dates = window["loan_dates"].values

        def _npv_at_target(adj):
            cf = np.concatenate([[-total_purchase], net_cf + adj])
            d = np.concatenate([[agg["loan_dates"].iloc[0]], dates])
            return xirr(cf.tolist(), d.tolist()) * 12 - irr_target_pct / 100

        try:
            from scipy.optimize import brentq
            add = brentq(_npv_at_target, -total_purchase, total_purchase, xtol=1e-6)
        except Exception:
            add = 0.0

        reserve_cumsum += add
        rows.append({
            "year": yr,
            "reserve_cost_cumsum": reserve_cumsum,
            "Reverse_Servicing_cost_addition": add,
        })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def run_pipeline(
    input_path: str,
    output_path: str | None = None,
    cprshock: float = 1.0,
    cdrshock: float = 1.0,
    irr_target: float = 7.9,
    workers: int | None = None,
    chunk_size: int | None = None,
    progress_callback: Callable[[int, str], None] | None = None,
    should_cancel: Callable[[], bool] | None = None,
) -> str:
    """Execute the full cashflow pipeline and write the Excel workbook.

    Parameters
    ----------
    input_path : str
        Path to current_assets.csv.
    output_path : str, optional
        Output .xlsx path. Defaults to <input_dir>/cashflows_<today>.xlsx.
    cprshock, cdrshock : float
        Shock multipliers (1.0 = no shock).
    irr_target : float
        Annual IRR target in percent for the IRR-support solver.
    workers : int, optional
        Number of worker processes for per-loan modeling. Defaults to 1 on
        Windows and auto-scales on larger Linux workloads.
    chunk_size : int, optional
        CSV rows to process per chunk. Defaults to a single chunk for smaller
        files and 5,000 rows for larger files.

    Returns
    -------
    str : resolved output path.
    """
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.parent / f"cashflows_{date.today():%Y%m%d}.xlsx"
    output_path = Path(output_path)

    print(f"Reading {input_path} …")
    if progress_callback:
        progress_callback(2, f"Reading {input_path.name}")
    n = _count_csv_rows(input_path)
    print(f"Processing {n} loans …")
    if progress_callback:
        progress_callback(5, f"Preparing {n:,} loans")

    zero_row = np.zeros(len(_CASHFLOW_COLS), dtype=float)
    bd_accum = defaultdict(lambda: zero_row.copy())
    non_bd_accum = defaultdict(lambda: zero_row.copy())
    prime_accum = defaultdict(lambda: zero_row.copy())
    sfy_accum = defaultdict(lambda: zero_row.copy())
    irr_accum = defaultdict(lambda: zero_row.copy())
    total_purchase = 0.0
    if workers is None:
        workers = 1 if sys.platform == "win32" or n < 5_000 else max(1, (os.cpu_count() or 2) - 1)
    else:
        workers = max(1, int(workers))
    if chunk_size is None:
        chunk_size = n if n <= 5_000 else 5_000
    else:
        chunk_size = max(1, int(chunk_size))

    executor = None
    if workers > 1:
        print(f"Using {workers} worker processes …")
        executor = ProcessPoolExecutor(max_workers=workers)

    with tempfile.TemporaryDirectory(prefix="cashflow-summaries-"):
        workbook = Workbook(write_only=True)
        bd_sheet = workbook.create_sheet(title="BD Cashflows")
        non_bd_sheet = workbook.create_sheet(title="Non-BD Cashflows")
        stack_upb_sheet = workbook.create_sheet(title="18 mth Stack")
        stack_mv_sheet = workbook.create_sheet(title="MV UPB 18 mth Stack")
        prime_sfy_mv_sheet = workbook.create_sheet(title="PRIME SFY MV Stack")
        prime_sheet = workbook.create_sheet(title="PRIME Data")
        sfy_sheet = workbook.create_sheet(title="SFY Data")
        irr_support_sheet = workbook.create_sheet(title="IRR Support")
        prime_sheet.append(_LOAN_SUMMARY_COLS)
        sfy_sheet.append(_LOAN_SUMMARY_COLS)
        successful_loans = 0
        try:
            processed = 0
            for chunk in pd.read_csv(input_path, low_memory=False, chunksize=chunk_size):
                if should_cancel and should_cancel():
                    raise InterruptedError("Cashflow job cancelled.")
                for col in _DATE_COLUMNS:
                    chunk[col] = pd.to_datetime(chunk[col])
                chunk["platform"] = chunk["platform"].str.lower()
                row_field_positions = {field: chunk.columns.get_loc(field) for field in _ROW_FIELDS}
                row_iter = (
                    ({field: row_values[pos] for field, pos in row_field_positions.items()}, cprshock, cdrshock)
                    for row_values in chunk.itertuples(index=False, name=None)
                )
                if workers == 1:
                    results_iter = map(_process_loan_task, row_iter)
                else:
                    results_iter = executor.map(_process_loan_task, row_iter, chunksize=128)

                for sfc_df, our_df, summary, warning in results_iter:
                    if should_cancel and should_cancel():
                        if executor is not None:
                            for proc in getattr(executor, "_processes", {}).values():
                                proc.terminate()
                            executor.shutdown(wait=False, cancel_futures=True)
                            executor = None
                        raise InterruptedError("Cashflow job cancelled.")
                    processed += 1
                    if processed % 500 == 0 or processed == n:
                        print(f"  {processed}/{n} ({processed/n*100:.0f}%) …")
                    if progress_callback and (processed == n or processed % max(250, n // 100 or 1) == 0):
                        progress = min(85, 5 + int((processed / max(n, 1)) * 70))
                        progress_callback(progress, f"Processed {processed:,} of {n:,} loans")
                    if warning:
                        print(f"  WARNING: {warning}")
                        continue
                    monthly_our = _monthly_rollup(our_df)
                    tagging = str(summary.get("tagging", "")).upper()
                    platform = str(summary.get("platform", "")).lower()
                    if tagging == "BD":
                        _accumulate_rollup(bd_accum, monthly_our)
                    elif tagging == "NON-BD":
                        _accumulate_rollup(non_bd_accum, monthly_our)
                    if platform == "prime":
                        _accumulate_rollup(prime_accum, monthly_our)
                        prime_sheet.append([_normalize_excel_value(summary.get(col)) for col in _LOAN_SUMMARY_COLS])
                    elif platform == "sfy":
                        _accumulate_rollup(sfy_accum, monthly_our)
                        sfy_sheet.append([_normalize_excel_value(summary.get(col)) for col in _LOAN_SUMMARY_COLS])
                    _accumulate_rollup(irr_accum, monthly_our)
                    successful_loans += 1
                    total_purchase += float(summary["purchase_price"])
        finally:
            if executor is not None:
                executor.shutdown()

        if not successful_loans:
            raise RuntimeError("No loans processed successfully.")

        print("Aggregating by segment …")
        if progress_callback:
            progress_callback(88, "Aggregating portfolio segments")

        # BD / Non-BD cashflows (OUR case)
        bd_agg = _segment_frame_from_accumulator(bd_accum, "BD")
        non_bd_agg = _segment_frame_from_accumulator(non_bd_accum, "Non_BD")

        # 18-month UPB stacks
        stack_upb = _build_upb_stack(bd_agg, non_bd_agg)
        stack_mv = _build_upb_stack(
            bd_agg, non_bd_agg,
            upb_col_bd="BD_opening_mv_upb",
            upb_col_non_bd="Non_BD_opening_mv_upb",
            label_bd="Pool BD MV UPB",
            label_non_bd="Pool Non BD MV UPB",
        )

        # PRIME & SFY MV UPB stack
        prime_agg = _segment_frame_from_accumulator(prime_accum, "PRIME")
        sfy_agg = _segment_frame_from_accumulator(sfy_accum, "SFY")
        prime_sfy_mv_stack = _build_upb_stack(
            prime_agg.rename(columns={"dates": "dates"}),
            sfy_agg.rename(columns={"dates": "dates"}),
            upb_col_bd="PRIME_opening_mv_upb",
            upb_col_non_bd="SFY_opening_mv_upb",
            label_bd="Pool PRIME MV UPB",
            label_non_bd="Pool SFY MV UPB",
        )

        # IRR support solver
        print("Solving IRR support adjustments …")
        if progress_callback:
            progress_callback(93, "Solving IRR support")
        try:
            irr_monthly_agg = pd.DataFrame(
                [{"dates": d, **dict(zip(_CASHFLOW_COLS, irr_accum[d]))} for d in sorted(irr_accum)]
            )
            irr_support_df = _solve_irr_support(irr_monthly_agg, total_purchase, irr_target)
        except Exception as exc:
            print(f"  WARNING: IRR support solver failed — {exc}")
            irr_support_df = pd.DataFrame(
                {"year": range(1, 11), "reserve_cost_cumsum": [0.0] * 10,
                 "Reverse_Servicing_cost_addition": [0.0] * 10}
            )

        print(f"Writing {output_path} …")
        if progress_callback:
            progress_callback(97, f"Writing {output_path.name}")
        bd_sheet.append(list(bd_agg.columns))
        _append_frame_rows(bd_sheet, bd_agg)
        non_bd_sheet.append(list(non_bd_agg.columns))
        _append_frame_rows(non_bd_sheet, non_bd_agg)
        stack_upb_sheet.append(list(stack_upb.columns))
        _append_frame_rows(stack_upb_sheet, stack_upb)
        stack_mv_sheet.append(list(stack_mv.columns))
        _append_frame_rows(stack_mv_sheet, stack_mv)
        prime_sfy_mv_sheet.append(list(prime_sfy_mv_stack.columns))
        _append_frame_rows(prime_sfy_mv_sheet, prime_sfy_mv_stack)
        irr_support_sheet.append(list(irr_support_df.columns))
        _append_frame_rows(irr_support_sheet, irr_support_df)
        workbook.save(output_path)

        print(f"Done. Output: {output_path}")
        if progress_callback:
            progress_callback(100, f"Finished {output_path.name}")
        return str(output_path)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def _parse_args(argv=None):
    p = argparse.ArgumentParser(
        description="Run behavioral cashflow model on current_assets.csv"
    )
    p.add_argument("--input", required=True, help="Path to current_assets.csv")
    p.add_argument("--output", default=None, help="Output .xlsx path (optional)")
    p.add_argument("--cprshock", type=float, default=1.0, help="CPR multiplier shock")
    p.add_argument("--cdrshock", type=float, default=1.0, help="CDR multiplier shock")
    p.add_argument("--target", type=float, default=7.9, help="IRR target in percent")
    p.add_argument("--workers", type=int, default=None, help="Worker processes for per-loan modeling")
    p.add_argument("--chunk-size", type=int, default=None, help="CSV rows per processing chunk")
    return p.parse_args(argv)


if __name__ == "__main__":
    args = _parse_args()
    run_pipeline(
        input_path=args.input,
        output_path=args.output,
        cprshock=args.cprshock,
        cdrshock=args.cdrshock,
        irr_target=args.target,
        workers=args.workers,
        chunk_size=args.chunk_size,
    )
